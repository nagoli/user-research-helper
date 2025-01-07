import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from typing import Dict, List
from user_research_helper.result_analysis.data import SegmentDataset

def create_excel_report(dataset: SegmentDataset, output_file: str):
    """
    Create an Excel report from segment dataset analysis results
    
    Args:
        dataset: SegmentDataset containing questions and segment answers
        output_file: Path to save the Excel report
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Segment Analysis Results"

    # Write headers
    ws.cell(row=1, column=1, value="Segment Name")
    for col, question in enumerate(dataset.questions, start=2):
        ws.cell(row=1, column=col, value=question.text)
        ws.cell(row=1, column=col).alignment = Alignment(wrap_text=True)

    # Define fill colors based on confidence
    high_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green
    medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
    low_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red

    # Process each segment in the dataset
    for row, (segment_name, segment_answers) in enumerate(dataset.segments.items(), start=2):
        # Write segment name
        ws.cell(row=row, column=1, value=segment_name)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        
        # Write analysis results for each question
        for col, question in enumerate(dataset.questions, start=2):
            cell = ws.cell(row=row, column=col)
            
            if question.id in segment_answers:
                answer = segment_answers[question.id]
                cell.value = answer.answer_summary
                
                # Set cell color based on confidence if available
                confidence = getattr(answer, 'confidence', 'low')
                if confidence == 'high':
                    cell.fill = high_fill
                elif confidence == 'medium':
                    cell.fill = medium_fill
                else:  # low confidence
                    cell.fill = low_fill
                
                cell.alignment = Alignment(wrap_text=True)

    # Set column widths
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 70

    # Save workbook
    wb.save(output_file)

def test_report_builder():
    from user_research_helper.result_analysis.data import Question, SegmentAnswer, SegmentDataset
    
    # Create test data
    dataset = SegmentDataset(
        questions=[
            Question(id="Q1", text="First question", column_index=0),
            Question(id="Q2", text="Second question", column_index=1),
        ],
        segments={
            "accessibility": {
                "Q1": SegmentAnswer(segment_name="accessibility", question_id="Q1", answer_summary="Easy to navigate", confidence="high"),
                "Q2": SegmentAnswer(segment_name="accessibility", question_id="Q2", answer_summary="Some keyboard issues", confidence="medium")
            },
            "performance": {
                "Q1": SegmentAnswer(segment_name="performance", question_id="Q1", answer_summary="Fast loading times", confidence="high"),
                "Q2": SegmentAnswer(segment_name="performance", question_id="Q2", answer_summary="Occasional lag", confidence="low")
            },
            "usability": {
                "Q1": SegmentAnswer(segment_name="usability", question_id="Q1", answer_summary="Intuitive interface", confidence="medium"),
                "Q2": SegmentAnswer(segment_name="usability", question_id="Q2", answer_summary="Clear workflow", confidence="high")
            }
        }
    )
    
    create_excel_report(dataset, "test_segment_report.xlsx")

if __name__ == "__main__":
    test_report_builder()
