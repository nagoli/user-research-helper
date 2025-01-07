import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
from typing import List
from user_research_helper.result_analysis.data import ResultAnalysis

def create_result_report(results: List[ResultAnalysis], output_file: str):
    """
    Create an Excel report from result analysis
    
    Args:
        results: List of ResultAnalysis containing questions and their analysis
        output_file: Path to save the Excel report
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Result Analysis"

    # Write headers
    ws.cell(row=1, column=1, value="Question")
    ws.cell(row=1, column=2, value="Analysis")
    ws.cell(row=1, column=1).alignment = Alignment(wrap_text=True)
    ws.cell(row=1, column=2).alignment = Alignment(wrap_text=True)

    # Define fill colors based on confidence
    
    medium_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Orange
    low_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")    # Red

    # Process each result
    for row, result in enumerate(results, start=2):
        # Write question text
        ws.cell(row=row, column=1, value=result.question_text)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True)
        
        # Write analysis with confidence-based color
        cell = ws.cell(row=row, column=2, value=result.analysis)
        cell.alignment = Alignment(wrap_text=True)
        
        # Set cell color based on confidence
        if result.confidence:
            if result.confidence == 'high':
                pass
            elif result.confidence == 'medium':
                cell.fill = medium_fill
            else:  # low confidence
                cell.fill = low_fill

    # Set column widths
    ws.column_dimensions['A'].width = 70  # Question column
    ws.column_dimensions['B'].width = 100  # Analysis column

    # Save workbook
    wb.save(output_file)

def test_result_report():
    from user_research_helper.result_analysis.data import ResultAnalysis
    
    # Create test data
    results = [
        ResultAnalysis(
            question_id="Q1",
            question_text="What is your experience with the product?",
            analysis="Overall positive experience across segments, with particular appreciation for the interface",
            confidence="high"
        ),
        ResultAnalysis(
            question_id="Q2",
            question_text="How easy was it to navigate?",
            analysis="Navigation was generally intuitive, though some mobile users reported minor issues",
            confidence="medium"
        )
    ]
    
    create_result_report(results, "test_result_report.xlsx")

if __name__ == "__main__":
    test_result_report()
